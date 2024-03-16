from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from main.models import Menu, Category

class Command(BaseCommand):
    help = 'Load data from myuni data into Django model'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='myuni-data.xlsx')

    def handle(self, *args, **options):
        file_path = options['file_path']
        wb = load_workbook(file_path)
        #sheet = wb.active
        drinks_category, _ = Category.objects.get_or_create(name='Drinks')
        desserts_category, _ = Category.objects.get_or_create(name='Desserts')
        maincourses_category, _ = Category.objects.get_or_create(name='Maincourses')
        
        drinks_sheet = wb['Drinks']
        for row in drinks_sheet.iter_rows(min_row=2, values_only=True):
            #category = Category.objects.get_or_create(name='Drinks')
            menu_item = Menu.objects.create(
                name=row[0],
                price=row[1],
                category=drinks_category
            )
            print(menu_item)
            
        desserts_sheet = wb['Desserts']
        for row in desserts_sheet.iter_rows(min_row=2, values_only=True):
            #category = Category.objects.get_or_create(name='Desserts')
            menu_item = Menu.objects.create(
                name=row[0],
                price=row[1],
                category=desserts_category
            )
            print(menu_item)
            
        maincourses_sheet = wb['Maincourses']
        for row in maincourses_sheet.iter_rows(min_row=2, values_only=True):
            #category = Category.objects.get_or_create(name='Maincourses')
            menu_item = Menu.objects.create(
                name=row[0],
                price=row[1],
                category=maincourses_category
            )
            print(menu_item)
            

        self.stdout.write(self.style.SUCCESS('Data loaded successfully'))